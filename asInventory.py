print ("Loading dependancies...")
import os
import sys
from archives_tools import aspace as AS
from archives_tools import uaLocations
import openpyxl
import traceback
import datetime
import shutil

def updateDate(fileObject, normal, display):
	if display.lower().strip() == "none":
		display = ""
	if "/" in normal:
		#date range
		if len(display) > 0:
			fileObject = AS.makeDate(fileObject, normal.split("/")[0], normal.split("/")[1], display)
		else:
			fileObject = AS.makeDate(fileObject, normal.split("/")[0], normal.split("/")[1])
	else:
		#single date
		if len(display) > 0:
			fileObject = AS.makeDate(fileObject, normal, "", display)
		else:
			fileObject = AS.makeDate(fileObject, normal)
	return fileObject

try:

	if sys.argv[1].lower().endswith("download"):
		outputPath = sys.argv[2]
		level = sys.argv[3]
		cmpntID = sys.argv[4]
		baseURL = sys.argv[5]
		repository = sys.argv[6]
		user = sys.argv[7]
		password = sys.argv[8]
		
		loginData = (baseURL, user, password)
		session = AS.getSession(loginData)
		sys.stdout.flush()
		
		#create Workbook object
		wb = openpyxl.Workbook()
		
		if level.lower().strip() == "resource":
			resourceLevel = True
			print ("Looking for resource")
		else:
			resourceLevel = False
			print ("Looking for  archival object")
		sys.stdout.flush()
		
		if resourceLevel == True:
			object = AS.getResourceID(session, repository, cmpntID, loginData)		
			displayTitle = object.title.replace("/", "-")
		else:	
			object = AS.getArchObjID(session, repository, cmpntID, loginData)
			displayTitle = object.display_string
			
		simpleTitle = object.title.replace("/", "-")
		print ("Reading " + simpleTitle)
		
		
		if resourceLevel == True:
			objectID = object.id_0
		else:
			objectID = object.ref_id
		
		#make a sheet
		worksheet = wb.active
		#worksheet = wb.create_sheet(title=simpleTitle)
		worksheet["H1"] = "Title"
		worksheet["I1"] = displayTitle
		worksheet["H2"] = "Level"
		if resourceLevel == True:
			worksheet["I2"] = "resource"
		else:
			worksheet["I2"] = "archival object"
		worksheet["H3"] = "Ref ID"
		worksheet["I3"] = objectID
		
		#setup table headings
		worksheet["A6"] = "ID"
		worksheet["B6"] = "Location ID"
		worksheet["C6"] = "Location"
		worksheet["D6"] = "Container URI"
		worksheet["E6"] = "Container"
		worksheet["F6"] = "C#"
		worksheet["G6"] = "Folder"
		worksheet["H6"] = "F#"
		worksheet["I6"] = "Title"
		worksheet["J6"] = "Date 1 Display"
		worksheet["K6"] = "Date 1 Normal"
		worksheet["L6"] = "Date 2 Display"
		worksheet["M6"] = "Date 2 Normal"
		worksheet["N6"] = "Date 3 Display"
		worksheet["O6"] = "Date 3 Normal"
		worksheet["P6"] = "Date 4 Display"
		worksheet["Q6"] = "Date 4 Normal"
		worksheet["R6"] = "Date 5 Display"
		worksheet["S6"] = "Date 5 Normal"
		worksheet["T6"] = "Restrictions"
		worksheet["U6"] = "General Note"
		worksheet["V6"] = "Scope"
		worksheet["W6"] = "DAO Filename"
		
		#get table styles
		tableStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
		
		if resourceLevel == True:
			resourceTree = AS.getTree(session, object, loginData)
			childrenList = resourceTree.children
		else:
			childTree = AS.getChildren(session, object, loginData)
			childrenList = childTree.children
			
		lineCount = 6
		for child in childrenList:
			childObject = AS.getArchObj(session, child.record_uri, loginData)
			lineCount = lineCount + 1
			worksheet["A" + str(lineCount)] = childObject.ref_id
			worksheet["I" + str(lineCount)] = childObject.title
			try:
				print ("	exporting " + childObject.title)
			except:
				print ("	exporting non-ascii file...")
			sys.stdout.flush()
			
			# containers and locations		 
			if len(childObject.instances) > 0:
				if "sub_container" in childObject.instances[0].keys():
					container = childObject.instances[0].sub_container
					containerURI = container.top_container.ref
					worksheet["D" + str(lineCount)] = containerURI
					if "type_2" in container.keys():
						worksheet["G" + str(lineCount)] = container.type_2
					if "indicator_2" in container.keys():
						worksheet["H" + str(lineCount)] = container.indicator_2
						
					containerObject = AS.getContainer(session, containerURI, loginData)
					worksheet["E" + str(lineCount)] = containerObject.type
					worksheet["F" + str(lineCount)] = containerObject.indicator
					
					locationCount = 0
					for location in containerObject.container_locations:
						locationCount = locationCount + 1
						locationObject = AS.getLocation(session, location.ref, loginData)
						if "area" in locationObject.keys():
							locationCoordinates = locationObject.area + "-" + locationObject.coordinate_1_indicator
						else:
							locationCoordinates = locationObject.room + "-" + locationObject.coordinate_1_indicator
						if "coordinate_2_indicator" in locationObject.keys():
							locationCoordinates = locationCoordinates + "-" + locationObject.coordinate_2_indicator
						if "coordinate_3_indicator" in locationObject.keys():
							locationCoordinates = locationCoordinates + "-" + locationObject.coordinate_3_indicator
						if locationCount < 2:
							worksheet["B" + str(lineCount)] = locationObject.uri
							worksheet["C" + str(lineCount)] = locationCoordinates
						else:
							worksheet["B" + str(lineCount)] = worksheet["B" + str(lineCount)].value + "; " + locationObject.uri
							worksheet["C" + str(lineCount)] = worksheet["C" + str(lineCount)].value + "; " + locationCoordinates					
					
			
			#dates
			dateCount = 0
			for date in childObject.dates:
				dateCount = dateCount + 1
				if dateCount == 1:
					displayCell = "J"
					normalCell = "K"
				elif dateCount == 2:
					displayCell = "L"
					normalCell = "M"
				elif dateCount == 3:
					displayCell = "N"
					normalCell = "O"
				elif dateCount == 4:
					displayCell = "P"
					normalCell = "Q"
				elif dateCount == 5:
					displayCell = "R"
					normalCell = "S"
				elif dateCount > 5:
					raise ValueError("ERROR more than 5 dates for " + "uri: " + childObject.uri + " ref_id: " + childObject.ref_id)
				if "end" in date.keys():
					worksheet[normalCell + str(lineCount)] = date.begin + "/" + date.end
				else:
					worksheet[normalCell + str(lineCount)] = date.begin
				if "expression" in date.keys():
					worksheet[displayCell + str(lineCount)] = date.expression
				if "certainty" in date.keys():
					if worksheet[displayCell + str(lineCount)].value is None:
						worksheet[displayCell + str(lineCount)] = date.certainty
					else:
						worksheet[displayCell + str(lineCount)] = date.certainty + " " + worksheet[displayCell + str(lineCount)].value
			
			for note in childObject.notes:
				if note.type == "accessrestrict":
					subCount = 0
					for subnote in note.subnotes:
						subCount = subCount + 1
						if subCount < 1:
							worksheet["T" + str(lineCount)] = worksheet["T" + str(lineCount)] + "; " +  subnote.content
						else:
							worksheet["T" + str(lineCount)] = subnote.content
				elif note.type == "odd":
					subCount = 0
					for subnote in note.subnotes:
						subCount = subCount + 1
						if subCount < 1:
							worksheet["U" + str(lineCount)] = worksheet["U" + str(lineCount)] + "; " +  subnote.content
						else:
							worksheet["U" + str(lineCount)] = subnote.content
				elif note.type == "scopecontent":
					subCount = 0
					for subnote in note.subnotes:
						subCount = subCount + 1
						if subCount < 1:
							worksheet["V" + str(lineCount)] = worksheet["V" + str(lineCount)] + "; " +  subnote.content
						else:
							worksheet["V" + str(lineCount)] = subnote.content
			
		print ("Writing spreadsheet " + simpleTitle + ".xlsx to " + outputPath)
		sys.stdout.flush()
		
		table = openpyxl.worksheet.table.Table(ref='A6:W' + str(lineCount), displayName='Inventory', tableStyleInfo=tableStyle)
		worksheet.add_table(table)
		
		#styles for sheet info on top
		worksheet["H1"].style = "Accent1"
		worksheet["H2"].style = "Accent1"
		worksheet["H3"].style = "Accent1"
		
		#set column widths
		worksheet.column_dimensions["I"].width = 60.0
		worksheet.column_dimensions["F"].width = 15.0
		worksheet.column_dimensions["J"].width = 15.0
		worksheet.column_dimensions["K"].width = 15.0
		
		
		wb.save(filename = os.path.join(outputPath, simpleTitle + ".xlsx"))
		print ("Export Complete!")
		sys.stdout.flush()
		
	elif sys.argv[1].lower().endswith("upload"):
		
		inputPath = sys.argv[2]
		daoPath = sys.argv[3]
		baseURL = sys.argv[4]
		repository = sys.argv[5]
		user = sys.argv[6]
		password = sys.argv[7]
		loginData = (baseURL, user, password)
		
		print ("Reading input directory...")
		sys.stdout.flush()
		for spreadFile in os.listdir(inputPath):
			
			#find and load sheets
			if spreadFile.endswith(".xlsx"):
				spreadsheet = os.path.join(inputPath, spreadFile)
				print ("Reading " + spreadFile)
				sys.stdout.flush()
				wb = openpyxl.load_workbook(filename=spreadsheet, read_only=True)
				
				#validate sheets
				for sheet in wb.worksheets:
					checkSwitch = True
					try:
						if sheet["H1"].value.lower().strip() != "title":
							checkSwitch = False
						elif sheet["H2"].value.lower().strip() != "level":
							checkSwitch = False
						elif sheet["H3"].value.lower().strip() != "ref id":
							checkSwitch = False
						elif sheet["J6"].value.lower().strip() != "date 1 display":
							checkSwitch = False
						elif sheet["D6"].value.lower().strip() != "container uri":
							checkSwitch = False
					except:
						print ("ERROR: incorrect sheet " + sheet.title + " in file " + spreadFile)
						
					if checkSwitch == False:
						print ("ERROR: incorrect sheet " + sheet.title + " in file " + spreadFile)
						sys.stdout.flush()
					else:
					
						#Read sheet info
						print ("Reading sheet: " + sheet.title)
						sys.stdout.flush()
						
						displayName = sheet["I1"].value
						level = sheet["I2"].value
						refID = sheet["I3"].value
						
						session = AS.getSession(loginData)
						sys.stdout.flush()
						
						if level.lower().strip() == "resource":
							resourceLevel = True
							print ("Looking for resource matching " + displayName + "...")
							sys.stdout.flush()
							object = AS.getResourceID(session, repository, refID, loginData)
							resourceURI = object.uri
							print ("Found " + object.title)
							sys.stdout.flush()
						else:
							resourceLevel = False
							print ("Looking for archival object matching " + displayName + "...")
							sys.stdout.flush()
							object = AS.getArchObjID(session, repository, refID, loginData)
							print ("Found " + object.title)
							resourceURI = object.resource.ref
							parentURI = object.uri
							sys.stdout.flush()
							
							
						boxSession = {}
						rowCount = 0
						for row in sheet.rows:
							rowCount = rowCount + 1
							if rowCount > 6:
								fileCount = rowCount - 6
								
								#make sure there is a title
								if not row[8].value is None:
								
									#make new file object if its new
									if row[0].value is None:
										fileObject = AS.makeArchObj()
										fileObject.level = "file"
										if resourceLevel == True:
											fileObject.resource = {"ref": resourceURI}
										else:
											fileObject.parent = {"ref": parentURI}
											fileObject.resource = {"ref": resourceURI}
									else:
										#else get existing object
										fileObject = AS.getArchObjID(session, repository, str(row[0].value).strip(), loginData)
										
									#set title and position
									fileObject.title = row[8].value
									#print (str(row[8].value) + " --> position " + str(fileCount) )
									fileObject.position = int(fileCount)
									#print (fileObject.position)
									#clear dates
									fileObject.dates = []
									
									#enter dates
									if not row[10].value is None:
										fileObject = updateDate(fileObject, str(row[10].value), str(row[9].value))
									if not row[12].value is None:
										fileObject = updateDate(fileObject, str(row[12].value), str(row[11].value))
									if not row[14].value is None:
										fileObject = updateDate(fileObject, str(row[14].value), str(row[13].value))
									if not row[16].value is None:
										fileObject = updateDate(fileObject, str(row[16].value), str(row[15].value))
									if not row[18].value is None:
										fileObject = updateDate(fileObject, str(row[18].value), str(row[17].value))
										
									#scope note
									if not row[21].value is None:
										newNotes = []
										for note in fileObject.notes:
											if note.type == "scopecontent":
												pass
											else:
												newNotes.append(note)
										fileObject.notes = newNotes
										fileObject = AS.makeMultiNote(fileObject, "scopecontent", str(row[21].value))
									#general note
									if not row[20].value is None:
										newNotes = []
										for note in fileObject.notes:
											if note.type == "odd":
												pass
											else:
												newNotes.append(note)
										fileObject.notes = newNotes
										fileObject = AS.makeMultiNote(fileObject, "odd", str(row[20].value))
									#access restrict note
									if not row[19].value is None:
										newNotes = []
										for note in fileObject.notes:
											if note.type == "accessrestrict":
												pass
											else:
												newNotes.append(note)
										fileObject.notes = newNotes
										fileObject.restrictions_apply = True
										fileObject = AS.makeMultiNote(fileObject, "accessrestrict", str(row[19].value))
									
									
									
									#containers
									if not row[4].value is None and not row[5].value is None:
										#if there is container info entered in spreadsheet
										if not row[3].value is None or str(row[4].value) + " " + str(row[5].value) in boxSession.keys():
											#existing container
											if row[3].value is None:
												boxUri = boxSession[str(row[4].value) + " " + str(row[5].value)]
											else:
												boxUri = str(row[3].value).strip()
											boxObject = AS.getContainer(session, boxUri, loginData)
											#look for existing box link
											foundBox = False
											newInstances = []
											for instance in fileObject.instances:
												if "sub_container" in instance.keys():
													if instance.sub_container.top_container.ref == boxUri:
														foundBox = True
														newInstances.append(instance)
													elif "digital_object" in instance.keys():
														newInstances.append(instance)
											fileObject.instances = newInstances
											if foundBox == False:
												#link to existing box
												fileObject = AS.addToContainer(session, fileObject, boxUri, None, None, loginData)
												
											#modify existing box
											for instance in fileObject.instances:
												if "sub_container" in instance.keys():
													if instance["sub_container"]["top_container"]["ref"] == boxUri:
														if not row[4].value is None:
															instance["container"]["type_1"] = str(row[4].value).strip()
															boxObject.type = str(row[4].value).strip()
														if not row[5].value is None:
															instance["container"]["indicator_1"] = str(row[5].value).strip()
															boxObject.indicator = str(row[5].value).strip()
														if not row[6].value is None:
															instance["container"]["type_2"] = str(row[6].value).strip()
															instance["sub_container"]["type_2"] = str(row[6].value).strip()
														if not row[7].value is None:
															instance["container"]["indicator_2"] = str(row[7].value).strip()
															instance["sub_container"]["indicator_2"] = str(row[7].value).strip()
											#add any restrictions to box
											if not row[19].value is None:
												boxObject.restricted = True
											
											
											#update locations
											if not row[1].value is None:
												for locationURI in str(row[1].value).split(";"):
													locTest = False
													for location in boxObject.container_locations:
														if location.ref == locationURI.strip():
															locTest = True
													if locTest == False:
														boxObject = AS.addToLocation(boxObject, locationURI)
											elif not row[2].value is None:
												#remove existing locations
												boxObject.container_locations = []
												
												#location, but without URI
												locCount = 0
												for locationSet in str(row[2].value).split(";"):
													locCount = locCount + 1
													if "(" in locationSet:
														coordinates = locationSet.split("(")[0].strip()
														locationNote = locationSet.split("(")[1].replace(")", "").strip()
													else:
														coordinates = locationSet.strip()
														locationNote = None
													
													coordList = uaLocations.location2ASpace(coordinates.strip(), locationNote)
													if coordList[1] is False:
														#single location
														locTitle = coordList[0]["Title"]
														locationURI = AS.findLocation(session, locTitle, loginData)
														if len(coordList[0]["Note"]) > 0:
															if locCount > 1:
																boxObject = AS.addToLocation(boxObject, locationURI, coordList[0]["Note"], "previous", "2999-01-01")
															else:
																boxObject = AS.addToLocation(boxObject, locationURI, coordList[0]["Note"])
														else:
															if locCount > 1:
																boxObject = AS.addToLocation(boxObject, locationURI, None, "previous", "2999-01-01")
															else:
																boxObject = AS.addToLocation(boxObject, locationURI)
													else:
														#multiple locations
														for location in coordList[0]:
															locTitle = location["Title"]
															locationURI = AS.findLocation(session, locTitle, loginData)
															if len(location["Note"]) > 0:
																if locCount > 1:
																	boxObject = AS.addToLocation(boxObject, locationURI, location["Note"], "previous", "2999-01-01")
																else:
																	boxObject = AS.addToLocation(boxObject, locationURI, location["Note"])
															else:
																if locCount > 1:
																	boxObject = AS.addToLocation(boxObject, locationURI, None, "previous", "2999-01-01")
																else:
																	boxObject = AS.addToLocation(boxObject, locationURI)
												print ("		Added location(s) to containers")
												sys.stdout.flush()
													
											
										else:
											#new box
											
											#delete existing boxes
											newInstances = []
											for instance in fileObject.instances:
												if "sub_container" in instance.keys():
													pass
												else:
													newInstances.append(instance)
											fileObject.instances = newInstances
											#makes and posts a new container
											boxObject = AS.makeContainer(session, repository, str(row[4].value), str(row[5].value), loginData)
											#update dict of new boxes for this sheet
											boxSession[str(row[4].value) + " " + str(row[5].value)] = boxObject.uri
											if not row[6].value is None:
												childContainer = str(row[6].value)
											else:
												childContainer = None
											if not row[7].value is None:
												childIndicator = str(row[7].value)
											else:
												childIndicator = None
											fileObject = AS.addToContainer(session, fileObject, boxObject.uri, childContainer,  childIndicator, loginData)
											#add any restrictions to box
											if not row[19].value is None:
												boxObject.restricted = True
												
											#update locations
											if not row[1].value is None:
												for locationURI in str(row[1].value).split(";"):
													locTest = False
													for location in boxObject.container_locations:
														if location.ref == locationURI.strip():
															locTest = True
													if locTest == False:
														boxObject = AS.addToLocation(boxObject, locationURI)
											elif not row[2].value is None:
												#remove existing locations from boxObject
												boxObject.container_locations = []
											
												#location, but without URI
												locCount = 0
												for locationSet in str(row[2].value).split(";"):
													locCount = locCount + 1
													if "(" in locationSet:
														coordinates = locationSet.split("(")[0].strip()
														locationNote = locationSet.split("(")[1].replace(")", "").strip()
													else:
														coordinates = locationSet.strip()
														locationNote = None
													
													coordList = uaLocations.location2ASpace(coordinates.strip(), locationNote)
													if coordList[1] is False:
														#single location
														locTitle = coordList[0]["Title"]
														locationURI = AS.findLocation(session, locTitle, loginData)
														if len(coordList[0]["Note"]) > 0:
															if locCount > 1:
																boxObject = AS.addToLocation(boxObject, locationURI, coordList[0]["Note"], "previous", "2999-01-01")
															else:
																boxObject = AS.addToLocation(boxObject, locationURI, coordList[0]["Note"])
														else:
															if locCount > 1:
																boxObject = AS.addToLocation(boxObject, locationURI, None, "previous", "2999-01-01")
															else:
																boxObject = AS.addToLocation(boxObject, locationURI)
													else:
														#multiple locations
														for location in coordList[0]:
															locTitle = location["Title"]
															locationURI = AS.findLocation(session, locTitle, loginData)
															if len(location["Note"]) > 0:
																if locCount > 1:
																	boxObject = AS.addToLocation(boxObject, locationURI, location["Note"], "previous", "2999-01-01")
																else:
																	boxObject = AS.addToLocation(boxObject, locationURI, location["Note"])
															else:
																if locCount > 1:
																	boxObject = AS.addToLocation(boxObject, locationURI, None, "previous", "2999-01-01")
																else:
																	boxObject = AS.addToLocation(boxObject, locationURI)
												print ("		Added location(s) to containers")
												sys.stdout.flush()
											
										#post top container object
										postBox = AS.postContainer(session, repository, boxObject, loginData)
										if postBox == 200:
											print ("		Posted " + str(row[4].value) + " " + str(row[5].value))
										else:
											print ("	Failed to post " +  str(row[4].value) + " " + str(row[5].value) + ", error code " + str(postBox))
											AS.pp(boxObject)
										sys.stdout.flush()
									
										
									
									#post file object
									postAO = AS.postArchObj(session, repository, fileObject, loginData)
									if postAO.status_code == 200:
										try:
											print ("	Posted " + row[8].value)
										except:
											print ("	Posted non-ascii text")
									else:
										print ("	Failed to post, error code " + str(postAO))
									sys.stdout.flush()
									
									#Digital Object
									if not row[22].value is None:
										#if post was successful
										if postAO.status_code == 200:
											print ("	-->Uploading dao for " + str(row[22].value))
											sys.stdout.flush()
										
											webDir = "\\\\romeo\\wwwroot\\eresources\\dao"
											webLink = "http://library.albany.edu/speccoll/findaids/eresources/dao"
											
											aoURI = postAO.json()["uri"]
											ao = AS.getArchObj(session, aoURI, loginData)
											aoRef = ao["ref_id"]
											
											resourceURI = ao["resource"]["ref"]
											coll = AS.getResource(session, repository, resourceURI.split("/resources/")[1], loginData)
											resourceID = coll["id_0"]
											uploadDir = os.path.join(webDir, resourceID)
											if not os.path.isdir(uploadDir):
												os.makedirs(uploadDir)
											uploadFile = os.path.join(daoPath, row[22].value)
											fileTitle = os.path.basename(row[22].value)
											extension = os.path.splitext(uploadFile)[1]
											daoLink = webLink + "/" + resourceID + "/" + aoRef + extension
											finalFile = os.path.join(uploadDir, aoRef + extension)
											
											#move file to web server
											shutil.move(uploadFile, finalFile)
											
											daoObject = AS.makeDAO(fileTitle, daoLink)
											postDAO = AS.postDAO(session, repository, daoObject, loginData)
											if postDAO.status_code == 200:
												daoURI = postDAO.json()["uri"]
												
												ao = AS.addDAO(ao, daoURI)
												postAO = AS.postArchObj(session, repository, ao, loginData)
												if not postAO.status_code == 200:
													raise ValueError("Error posting archival object with digital object " + row[22].value)
																								
											else:
												raise ValueError("Error posting digital object " + row[22].value)
																					
										
									
				wb._archive.close()
				print ("Moving " + spreadFile + " to complete directory...")
				sys.stdout.flush()
				completeDir = os.path.join(os.path.dirname(inputPath), "complete")
				if os.path.isfile(os.path.join(completeDir, spreadFile)):
					shutil.copy2(os.path.join(inputPath, spreadFile), os.path.join(completeDir, os.path.splitext(spreadFile)[0] + str(datetime.datetime.now()).split(".")[0].replace(":", "_") + ".xlsx"))		
				else:
					shutil.copy2(os.path.join(inputPath, spreadFile), completeDir)
				os.remove(os.path.join(inputPath, spreadFile))
			else:
				print ("ERROR: incorrect file " + spreadFile + " in input path.")
		
		
		
	else:
		raise ValueError("ERROR: recieved incorrect args: \n" + sys.argv)
		
except:
	exceptMsg = traceback.format_exc()
	print (exceptMsg)
	errorOutput = "\n" + "#############################################################\n" + str(datetime.datetime.now()) + "\n#############################################################\n" + str(exceptMsg) + "\n********************************************************************************"
	__location__ = os.path.dirname(os.path.abspath(__file__))
	file = open(os.path.join(__location__, "error.log"), "a")
	file.write(errorOutput)
	file.close()
	sys.exit(3)	