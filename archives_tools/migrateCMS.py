import aspace as AS
import datetime
from openpyxl import load_workbook
import os

if os.name == "nt":
	#Windows Directory Names

	#Collection and Subject spreadsheets directory
	spreadDir = "\\\\romeo\\Collect\spe\\Tools\\CollectionList"
	cmsDir = "\\\\romeo\\Collect\\spe\\Greg\\CMSdata"
	
else:
	#Unix directory names

	#Collection and Subject spreadsheets directory
	spreadDir = "/media/bcadmin/Collect/spe/Tools/CollectionList"
	

#parse Collection List spreadsheet
collectionListFile = os.path.join(spreadDir, "collectionList.xlsx")
collectionWorkbook = load_workbook(filename = collectionListFile, use_iterators=True, read_only=True)
collectionList = collectionWorkbook.get_sheet_by_name('collectionList')

#parse Local Subject Lists spreadsheet
subjectGuidesFile =  os.path.join(spreadDir, "subjectGuides.xlsx")
subjectWorkbook = load_workbook(filename = subjectGuidesFile, use_iterators=True, read_only=True)
subjectGuides = subjectWorkbook.get_sheet_by_name('subjectGuides')


#Parse List of Collections to list of lists
rowIndex = 0
collections = []
for row in collectionList.rows:
	rowIndex = rowIndex + 1
	if rowIndex > 1:
		collection = [str(rowIndex), row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value, row[10].value, row[11].value]
		collections.append(collection)
		
session = AS.getSession()
repo = "2"

print ("Getting List of IDs from ASpace...")
resourceIDs = []
for resource in AS.getResources(session, repo, "all"):
	resourceID = resource.ead_id
	if "_" in resourceID:
		resourceID = resourceID.split("_")[1]
	resourceIDs.append(resourceID)
	
print ("Getting a list of subjects from ASpace...")
subjectData = AS.getSubjects(session, "all")

	
for collection in collections:
	if str(collection[6]).lower().strip() == "null" or str(collection[6]).lower().strip() == "undated":
		print ("No Date for " +  collection[1] + " " +  collection[4])
	elif collection[8].lower().strip() == "null":
		print ("No Extent for " +  collection[1] + " " +  collection[4])
	else:
		if not collection[1] in resourceIDs:
			print ("making resouce for " + collection[4])
		
			newRes = AS.makeResource()
			resType = collection[5]
			collectionName = collection[4]
			if "," in collectionName and collection[5].lower().strip() == "papers":
				fixedName = collectionName.split(",")[1].strip() + " " + collectionName.split(",")[0].strip()
			elif collectionName.lower().strip().startswith("office of"):
				fixedName = collectionName.split(",")[1].strip() + " " + collectionName.split(",")[0].strip()
			else:
				fixedName = collectionName
			
			newRes.title = fixedName + " " + resType
			newRes.level = "collection"
			colID = collection[1].strip()
			typeList = ["collection", "papers", "publications", "records"]
			if resType.lower() in typeList:
				newRes.resource_type = resType.lower()
			else:
				newRes.resource_type = "collection"
			
			restrict = collection[2]
			if not restrict is None:
				newRes.publish = False
				newRes = AS.makeMultiNote(newRes, "accessrestrict", restrict)
			
			for note in newRes.notes:
				if note["type"] == "accessrestrict":
					note["publish"] = True
					for subnote in note["subnotes"]:
						subnote["publish"] = True
				
			else:
				newRes.publish = True
				
			newRes.id_0 = colID
			newRes.ead_id = colID
			newRes.ead_location = "http://meg.library.albany.edu:8080/archive/view?docId=" + colID + ".xml"

			if collection[3].lower().strip() == "html":
				newRes.finding_aid_note = "HTML Container List"

			newRes.finding_aid_author = "Migrated from CMS and Drupal Abstracts"
			newRes.finding_aid_date = datetime.datetime.now().isoformat().split("T")[0]


			#newRes = newRes.AS.addContainerLocation(newRes, "Collection", locationList, locationNote)
			
			dateField = str(collection[6])
			if "ca. " in dateField.lower().strip():
				dateField = dateField.replace("ca. ", "")
			if not dateField.lower().strip() == "null":
				if "," in dateField:
					for dateText in dateField.split(","):
						date = dateText.strip()
						if "-" in date:
							newRes = AS.makeDate(newRes, date.split("-")[0], date.split("-")[1])
						else:
							newRes = AS.makeDate(newRes, date.split("-")[0], "")
				else:
					if "-" in dateField:
						newRes = AS.makeDate(newRes, dateField.split("-")[0], dateField.split("-")[1])
					else:
						newRes = AS.makeDate(newRes, dateField.split("-")[0], "")
					
			
			newRes = AS.makeExtent(newRes, collection[7], collection[8])
			
			abstractText = collection[9]
			newRes = AS.makeSingleNote(newRes, "abstract", abstractText)
			
			print ("	adding subjects")
			megList = []
			subjectSet = []
			subjectRowNumber = 0
			for subjectRow in subjectGuides.rows:
				subjectRowNumber = subjectRowNumber + 1
				if subjectRowNumber == 1:
					subjectSet = subjectRow
				else:
					
					colCount = 0			
					for subID in subjectRow:
						colCount = colCount + 1
						if not subID.value is None:
							if subID.value.lower().strip() == colID:
								megList.append(subjectSet[colCount - 1])
			
			for megSubject in megList:
				subjectFound = False
				for subject in subjectData:
					for entry in subject.terms:
						try:
							if entry.term == megSubject:
								subjectFound = True
								subjectRef = subject.uri
						except:
							pass
				
				if subjectFound == True:
					newRes = AS.addSubject(session, newRes, subjectRef)
							

			AS.postResource(session, repo, newRes) #2035